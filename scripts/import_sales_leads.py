"""Imports Brent's real "Sales Lead Tracker.xlsx" into the leads table, so
that spreadsheet can be retired in favor of the app's Leads page.

Only the "Sales Lead Tracker" sheet is imported. The workbook's "People to
remember" sheet is a different concept entirely -- referral-network contacts
(realtors, investors) to stay in touch with, not active sales opportunities
-- and is deliberately NOT imported here. Ask Brent whether that should
become its own thing later; not folded into leads by guesswork.

Skips (never overwrites) any lead that already matches an existing one by
(first_name, project_address) case-insensitively, so this is safe to re-run
after partial progress.

Run this yourself. Requires SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY set in
your shell.

Usage:
    cd api && source .venv/bin/activate
    python ../scripts/import_sales_leads.py "/path/to/Mud & Marble Sales Lead Tracker.xlsx"
"""

import os
import sys

import httpx
import openpyxl

# Exact labels from the workbook's own "Sales Stages" sheet, mapped onto this
# app's status vocabulary (Leads.tsx's STAGE_OPTIONS). Matched by prefix
# ("Stage 3" etc.) so trailing whitespace/punctuation in the real cells
# doesn't cause a miss. "Lost" (used once, on the Katelyn Prentice row,
# instead of the canonical "Stage 6: Missed Opportunity" label) maps to the
# same terminal state.
STAGE_MAP = {
    "stage 1": "stage_1",
    "stage 2": "stage_2",
    "stage 3": "stage_3",
    "stage 4": "stage_4",
    "stage 5": "stage_5",
    "stage 6": "stage_6_lost",
    "lost": "stage_6_lost",
}


def map_stage(raw) -> str:
    if not raw:
        return "new"
    text = str(raw).strip().lower()
    for prefix, status in STAGE_MAP.items():
        if text.startswith(prefix) or text == prefix:
            return status
    return "new"


def to_profit_and_note(raw) -> tuple:
    """Projected Profit is almost always a number, but a few real rows have a
    text note instead (e.g. "Estimate Due by EOD Friday") -- never discard
    that, route it into notes instead of silently dropping it."""
    if raw is None or raw == "":
        return None, None
    if isinstance(raw, (int, float)):
        return float(raw), None
    return None, str(raw).strip()


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python import_sales_leads.py '/path/to/Mud & Marble Sales Lead Tracker.xlsx'", file=sys.stderr)
        sys.exit(1)
    path = sys.argv[1]

    supabase_url = os.environ.get("SUPABASE_URL")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        print("Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY in your shell first.", file=sys.stderr)
        sys.exit(1)

    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Sales Lead Tracker" not in wb.sheetnames:
        print(f"No 'Sales Lead Tracker' sheet found in {path}. Sheets present: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    ws = wb["Sales Lead Tracker"]

    existing = httpx.get(f"{supabase_url}/rest/v1/leads?select=first_name,project_address", headers=headers)
    if not existing.is_success:
        print(f"Failed to read existing leads: {existing.status_code} {existing.text}", file=sys.stderr)
        sys.exit(1)
    existing_keys = {
        (str(l.get("first_name") or "").strip().lower(), str(l.get("project_address") or "").strip().lower())
        for l in existing.json()
    }

    created, skipped, errors = 0, 0, 0
    # Header lives on row 2 (row 1 is the "SALES LEAD TRACKER / TOTAL
    # PROJECTIONS" banner); real rows start at row 3.
    for row in ws.iter_rows(min_row=3):
        lead_name = row[1].value  # column B
        if not lead_name or not str(lead_name).strip():
            continue
        address = row[2].value  # C
        scope = row[3].value  # D
        profit_raw = row[4].value  # E
        # F (Lead Temp) and G (Last Engaged) are present as headers in the
        # sheet but empty on every real row as of this import -- included
        # here so a future re-run picks them up if Brent starts filling
        # them in, without needing to touch this script again.
        temp_raw = row[5].value  # F
        last_engaged_raw = row[6].value  # G
        stage_raw = row[7].value  # H
        notes_raw = row[8].value  # I
        objections_raw = row[9].value  # J

        first_name = str(lead_name).strip()
        proj_address = str(address).strip() if address else None
        key = (first_name.lower(), (proj_address or "").lower())
        if key in existing_keys:
            skipped += 1
            continue

        profit, profit_note = to_profit_and_note(profit_raw)
        notes_parts = [str(notes_raw).strip() if notes_raw else None, profit_note]
        notes = "\n".join(p for p in notes_parts if p) or None

        payload = {
            "first_name": first_name,
            "project_address": proj_address,
            "project_scope": str(scope).strip() if scope else None,
            "projected_profit": profit,
            "lead_temp": str(temp_raw).strip().lower() if temp_raw else None,
            "last_contacted_at": str(last_engaged_raw) if last_engaged_raw else None,
            "status": map_stage(stage_raw),
            "notes": notes,
            "objections": str(objections_raw).strip() if objections_raw else None,
        }
        payload = {k: v for k, v in payload.items() if v is not None}

        r = httpx.post(f"{supabase_url}/rest/v1/leads", headers=headers, json=payload)
        if not r.is_success:
            print(f"Failed to create '{first_name}': {r.status_code} {r.text}", file=sys.stderr)
            errors += 1
            continue
        created += 1
        print(f"Created: {first_name} ({proj_address or 'no address'})")

    print(f"\nDone. Created {created}, skipped {skipped} (already existed), {errors} errors.")
    print(
        "\nNote: the workbook's 'People to remember' sheet (referral-network contacts like "
        "realtors and investors) was NOT imported -- that's a different concept from active "
        "sales leads. Ask if you want that brought in separately."
    )


if __name__ == "__main__":
    main()
