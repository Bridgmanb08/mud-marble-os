"""One-time import: pulls transactions, subcontractor contracts/payments, and estimate
line items out of Brent's hand-maintained "In-House <Job Name>.xlsx" workbooks
(~/Downloads/10. GC Projects/) and into the matching project in the app.

Talks to the *real deployed app* (not Supabase directly), logging in with your own
account -- same pattern as import_estimates_from_excel.py. All the builder-cost/markup
math and cost-code/subcontractor matching goes through the same code the app itself
uses. Your password only ever goes to your own login endpoint.

Each workbook has 4 sheets:
  - Overview      -- all derived (formulas / blank manual balances), nothing to import.
  - Quickbooks    -- raw transactions under section headers (Income, GC Rehab, etc.),
                     ending in a computed "ESTIMATE VS COST" pivot (not imported).
  - Estimate      -- full line-item budget, imported as v1 estimate line items.
  - Contractors   -- repeating per-sub blocks: contract line items + payments.

Safe to re-run: estimate line items dedupe on (title, unit_cost); transactions dedupe on
(date, amount, description); subcontractor-items are NOT deduped against re-runs (there's
no natural key), so don't run this twice against a job whose Contractors sheet already
landed -- the script prints a summary each time so you can tell.

Requires: pip install openpyxl httpx

Usage:
    cd api && source .venv/bin/activate
    python ../scripts/import_inhouse_sheets.py "/Users/brentbridgman/Downloads/10. GC Projects"
"""

import getpass
import sys
from pathlib import Path

import httpx
import openpyxl

DEFAULT_BASE_URL = "https://project-n43cv.vercel.app"


def bucket_for(category: str, title: str) -> str:
    if (category or "").strip().upper().startswith("20 -"):
        return "allowance"
    if (title or "").strip().lower() in {"pm fee", "project management fee"}:
        return "pm_fee"
    return "construction"


def parse_cost_code(raw: str):
    if not raw:
        return None
    raw = str(raw)
    if " - " in raw:
        return raw.split(" - ", 1)[0].strip()
    return raw.strip()


def job_name_from_filename(path: Path) -> str:
    name = path.stem
    if name.lower().startswith("in-house "):
        name = name[len("in-house ") :]
    return name.strip()


STREET_SUFFIXES = {"st", "street", "ave", "avenue", "ln", "lane", "rd", "road", "dr", "drive", "blvd"}


def _strip_street_suffix(name: str) -> str:
    words = name.split()
    if len(words) > 1 and words[-1] in STREET_SUFFIXES:
        return " ".join(words[:-1])
    return name


def match_project(job_name: str, projects: list[dict]):
    needle = job_name.lower()
    needle_no_suffix = _strip_street_suffix(needle)
    for p in projects:
        prefix = p["name"].split("|")[0].strip().lower()
        prefix_no_suffix = _strip_street_suffix(prefix)
        if prefix in (needle, needle_no_suffix) or prefix_no_suffix in (needle, needle_no_suffix):
            return p, None
    matches = [p for p in projects if needle in p["name"].lower() or needle_no_suffix in p["name"].lower()]
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        return None, f"{len(matches)} ambiguous matches: " + ", ".join(m["name"] for m in matches)
    return None, "no matching project found"


def import_estimate_sheet(client: httpx.Client, project: dict, ws, codes_by_code: dict, unmatched: list):
    existing_estimates = client.get("/api/estimates", params={"project_id": project["id"]}).json()
    if existing_estimates:
        estimate = existing_estimates[0]
        existing_items = client.get(f"/api/estimates/{estimate['id']}/items").json()
        existing_keys = {(i["title"], i["unit_cost"]) for i in existing_items}
    else:
        est_r = client.post(
            "/api/estimates", json={"project_id": project["id"], "version": 1, "status": "draft", "pm_fee_total": 0}
        )
        if not est_r.is_success:
            print(f"    FAILED to create estimate: {est_r.status_code} {est_r.text}", file=sys.stderr)
            return 0, 0
        estimate = est_r.json()
        existing_keys = set()

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    col = {cell.value: i for i, cell in enumerate(header_row) if cell.value}

    def cell(row, name):
        i = col.get(name)
        if i is None or i >= len(row):
            return None
        v = row[i].value
        return v if v not in ("", None) else None

    created, already_present = 0, 0
    for row in ws.iter_rows(min_row=2):
        title = cell(row, "Title")
        if not title or not isinstance(title, str):
            continue
        try:
            unit_cost = float(cell(row, "Unit Cost") or 0)
            quantity = float(cell(row, "Quantity") or 1)
            markup_raw = float(cell(row, "Markup") or 0)
        except (TypeError, ValueError):
            print(f"    SKIPPED unexpected row for '{title}' (non-numeric cost/quantity/markup)", file=sys.stderr)
            continue
        if (title, unit_cost) in existing_keys:
            already_present += 1
            continue
        category = cell(row, "Category") or ""
        raw_cost_code = cell(row, "Cost Code") or ""
        markup_type_raw = cell(row, "Markup Type") or "%"
        description = cell(row, "Description")
        internal_notes = cell(row, "Internal Notes")

        markup_type, markup_value = ("percent", markup_raw) if markup_type_raw == "%" else ("flat", markup_raw)

        code = parse_cost_code(raw_cost_code)
        cost_code_id = codes_by_code.get(code.lower()) if code else None
        notes_internal = internal_notes or ""
        if raw_cost_code and not cost_code_id:
            notes_internal = f"{notes_internal}\n[Source cost code not matched: {raw_cost_code}]".strip()
            unmatched.append((project["name"], title, raw_cost_code))

        item_r = client.post(
            f"/api/estimates/{estimate['id']}/items",
            json={
                "cost_code_id": cost_code_id,
                "bucket": bucket_for(category, title),
                "title": title,
                "description": description,
                "quantity": quantity,
                "unit_cost": unit_cost,
                "cost_type": "none",
                "markup_type": markup_type,
                "markup_value": markup_value,
                "notes_external": description,
                "notes_internal": notes_internal or None,
            },
        )
        if not item_r.is_success:
            print(f"    FAILED line item '{title}': {item_r.status_code} {item_r.text}", file=sys.stderr)
            continue
        created += 1
    return created, already_present


def import_quickbooks_sheet(client: httpx.Client, project: dict, ws, codes_by_code: dict, existing_tx_keys: set):
    created, skipped = 0, 0
    section = None
    for row in ws.iter_rows(min_row=1):
        values = [c.value for c in row]
        non_none = [v for v in values if v not in (None, "")]
        # a lone text value in column B (index 1) marks a new section header
        if len(non_none) == 1 and isinstance(values[1], str) and values[1] not in ("QUICKBOOKS EXPORT",):
            if values[1] == "ESTIMATE VS COST":
                break  # everything after this is a computed pivot, not raw data
            section = values[1]
            continue
        date_val, name_val, memo_val, amount_val, code_val = (values[1:6] + [None] * 5)[:5]
        # Real transaction rows always have an actual date cell in this column.
        # Repeated column-header rows ("Date"), subtotal rows ("Total for ...",
        # "Net Income", etc.), and anything else that isn't a genuine dated
        # transaction all have plain text here instead -- skip them uniformly
        # rather than trying to enumerate every label that might show up.
        if not hasattr(date_val, "date") or not isinstance(amount_val, (int, float)):
            continue
        tx_date = date_val.date().isoformat()
        amount = float(amount_val)
        is_income = section == "Income"
        signed_amount = abs(amount) if is_income else -abs(amount)
        description = memo_val or None
        key = (tx_date, round(signed_amount, 2), (description or "")[:60])
        if key in existing_tx_keys:
            skipped += 1
            continue
        code = parse_cost_code(code_val) if code_val else None
        cost_code_id = codes_by_code.get(code.lower()) if code else None
        tx_r = client.post(
            "/api/transactions",
            json={
                "project_id": project["id"],
                "transaction_date": tx_date,
                "vendor": name_val or None,
                "transaction_type": "income" if is_income else "expense",
                "amount": signed_amount,
                "payment_source": section,
                "cost_code_id": cost_code_id,
                "description": description,
            },
        )
        if not tx_r.is_success:
            print(f"    FAILED transaction on {tx_date}: {tx_r.status_code} {tx_r.text}", file=sys.stderr)
            continue
        existing_tx_keys.add(key)
        created += 1
    return created, skipped


def find_or_create_subcontractor(client: httpx.Client, name: str, subs_by_name: dict):
    key = name.strip().lower()
    if key in subs_by_name:
        return subs_by_name[key]
    r = client.post("/api/subcontractors", json={"company_name": name.strip()})
    if not r.is_success:
        print(f"    FAILED to create subcontractor '{name}': {r.status_code} {r.text}", file=sys.stderr)
        return None
    sub = r.json()
    subs_by_name[key] = sub
    return sub


def import_contractors_sheet(client: httpx.Client, project: dict, ws, subs_by_name: dict, existing_tx_keys: set):
    items_created, payments_created, subs_touched = 0, 0, 0
    rows = list(ws.iter_rows(min_row=1))
    i = 0
    while i < len(rows):
        row = rows[i]
        values = [c.value for c in row]
        non_none = [v for v in values if v not in (None, "")]
        is_header_or_label = values[1] in (
            None,
            "CONTRACTORS",
            "Contract/Agreement",
            "Line Item",
            "Total Contract",
        )
        if len(non_none) == 1 and isinstance(values[1], str) and not is_header_or_label:
            name = values[1].strip()
            i += 1
            if name.upper() == "NAME":
                # unused placeholder block -- skip to its "Total Contract" row
                while i < len(rows) and rows[i][1].value != "Total Contract":
                    i += 1
                i += 1
                continue

            sub = find_or_create_subcontractor(client, name, subs_by_name)
            if sub is None:
                while i < len(rows) and rows[i][1].value != "Total Contract":
                    i += 1
                i += 1
                continue
            subs_touched += 1

            # skip the "Contract/Agreement | Payments" and "Line Item | Amount | Date | Amount | Category" header rows
            while i < len(rows) and rows[i][1].value in ("Contract/Agreement", "Line Item"):
                i += 1

            while i < len(rows) and rows[i][1].value != "Total Contract":
                r = rows[i]
                r_vals = [c.value for c in r]
                if r_vals[1] in ("Contract/Agreement", "Line Item"):
                    # a header row repeated mid-block, not a real entry
                    i += 1
                    continue
                desc, amount = r_vals[1] if len(r_vals) > 1 else None, r_vals[2] if len(r_vals) > 2 else None
                pay_date, pay_amount, pay_category = (
                    r_vals[4] if len(r_vals) > 4 else None,
                    r_vals[5] if len(r_vals) > 5 else None,
                    r_vals[6] if len(r_vals) > 6 else None,
                )
                if isinstance(amount, (int, float)):
                    item_r = client.post(
                        f"/api/projects/{project['id']}/subcontractor-items",
                        json={"subcontractor_id": sub["id"], "description": desc, "amount": float(amount)},
                    )
                    if item_r.is_success:
                        items_created += 1
                    else:
                        print(f"    FAILED contract item '{desc}': {item_r.status_code} {item_r.text}", file=sys.stderr)
                if hasattr(pay_date, "date") and isinstance(pay_amount, (int, float)):
                    tx_date = pay_date.date().isoformat()
                    if tx_date:
                        signed = -abs(float(pay_amount))
                        key = (tx_date, round(signed, 2), f"Payment to {name}"[:60])
                        if key not in existing_tx_keys:
                            tx_r = client.post(
                                "/api/transactions",
                                json={
                                    "project_id": project["id"],
                                    "transaction_date": tx_date,
                                    "vendor": name,
                                    "transaction_type": "expense",
                                    "amount": signed,
                                    "payment_source": pay_category,
                                    "subcontractor_id": sub["id"],
                                    "description": f"Payment to {name}",
                                },
                            )
                            if tx_r.is_success:
                                existing_tx_keys.add(key)
                                payments_created += 1
                            else:
                                print(f"    FAILED payment on {tx_date}: {tx_r.status_code} {tx_r.text}", file=sys.stderr)
                i += 1
            i += 1  # past "Total Contract" row
        else:
            i += 1
    return items_created, payments_created, subs_touched


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python import_inhouse_sheets.py <folder containing 'In-House *.xlsx' files>", file=sys.stderr)
        sys.exit(1)
    folder = Path(sys.argv[1])
    files = sorted(
        p for p in folder.glob("In-House *.xlsx") if p.name.lower() != "in-house template.xlsx"
    )
    if not files:
        print(f"No 'In-House *.xlsx' files found in {folder}", file=sys.stderr)
        sys.exit(1)

    base_url = input(f"App base URL [{DEFAULT_BASE_URL}]: ").strip() or DEFAULT_BASE_URL
    email = input("Email: ").strip()
    password = getpass.getpass("Password: ")

    client = httpx.Client(base_url=base_url, timeout=30)
    login = client.post("/api/auth/login", json={"email": email, "password": password})
    if not login.is_success:
        print(f"Login failed: {login.status_code} {login.text}", file=sys.stderr)
        sys.exit(1)
    print(f"Logged in as {login.json().get('email')}")

    cost_codes = client.get("/api/transactions/cost-codes").json()
    codes_by_code = {c["code"].strip().lower(): c["id"] for c in cost_codes if c.get("code")}
    subs = client.get("/api/subcontractors").json()
    subs_by_name = {s["company_name"].strip().lower(): s for s in subs}
    projects = client.get("/api/projects", params={"include_archived": "true"}).json()

    unmatched_codes = []
    skipped_files = []

    for path in files:
        job_name = job_name_from_filename(path)
        print(f"\n=== {path.name} -> job '{job_name}' ===")
        project, err = match_project(job_name, projects)
        if project is None:
            print(f"  SKIPPED: {err}")
            skipped_files.append((path.name, err))
            continue
        print(f"  Matched project: {project['name']} (id={project['id']})")

        try:
            wb = openpyxl.load_workbook(path, data_only=True)

            if "Estimate" in wb.sheetnames:
                created, present = import_estimate_sheet(client, project, wb["Estimate"], codes_by_code, unmatched_codes)
                note = f" ({present} already present, skipped)" if present else ""
                print(f"  Estimate: imported {created} line items.{note}")

            existing_tx = client.get("/api/transactions", params={"project_id": project["id"]}).json()
            existing_tx_keys = {
                (t["transaction_date"][:10], round(t["amount"], 2), (t.get("description") or "")[:60])
                for t in existing_tx
            }

            if "Quickbooks" in wb.sheetnames:
                created, skipped = import_quickbooks_sheet(
                    client, project, wb["Quickbooks"], codes_by_code, existing_tx_keys
                )
                note = f" ({skipped} already present, skipped)" if skipped else ""
                print(f"  Quickbooks: imported {created} transactions.{note}")

            if "Contractors" in wb.sheetnames:
                items, payments, subs_touched = import_contractors_sheet(
                    client, project, wb["Contractors"], subs_by_name, existing_tx_keys
                )
                print(f"  Contractors: {subs_touched} subcontractor(s), {items} contract items, {payments} payments.")
        except Exception as exc:  # noqa: BLE001 -- a surprise in one file must not block the rest
            print(f"  FAILED partway through this file: {exc!r}", file=sys.stderr)
            print("  Whatever imported before the error is already saved; re-run once fixed to pick up the rest.")
            continue

    if unmatched_codes:
        print("\n=== Line items whose source cost code didn't match your existing cost code list (noted on the item, not blocking) ===")
        for job_name, title, raw in unmatched_codes:
            print(f"  {job_name} / {title}: '{raw}'")

    if skipped_files:
        print("\n=== Files skipped -- couldn't confidently match a project ===")
        for name, err in skipped_files:
            print(f"  {name}: {err}")

    print("\nDone. Review each job's workshop (In-House Sheet -> job) before relying on the numbers.")


if __name__ == "__main__":
    main()
