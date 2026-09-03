import io
from datetime import datetime
from typing import Optional
from xml.sax.saxutils import escape as _xml_escape

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import HRFlowable, Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .. import branding
from ..deps import CurrentUser, get_current_user
from ..estimate_defaults import DEFAULT_CLOSING_TEXT
from ..estimate_text_defaults_store import get_or_create_estimate_text_defaults
from ..rich_text import rich_text_to_pdf_markup
from ..schemas.estimates import (
    EstimateCreate,
    EstimateOut,
    EstimateUpdate,
    LineItemCreate,
    LineItemOut,
    LineItemReference,
    LineItemUpdate,
)
from ..supabase_client import db_delete, db_get, db_patch, db_post, db_post_many

router = APIRouter(prefix="/estimates", tags=["estimates"])


def _compute_costs(quantity: float, unit_cost: float, markup_type: str, markup_value: float) -> tuple[float, float]:
    builder_cost = round((quantity or 0) * (unit_cost or 0), 2)
    if markup_type == "flat":
        owner_price = round(builder_cost + (markup_value or 0), 2)
    else:
        owner_price = round(builder_cost * (1 + (markup_value or 0) / 100), 2)
    return builder_cost, owner_price


async def _check_not_below_invoiced(item_id: str, new_owner_price: float) -> None:
    """Reducing a line item's price below what's already been invoiced
    against it (a real workflow -- price corrections happen after partial
    invoicing) had no guard at all: the invoice picker's remaining_amount
    (owner_price - invoiced_amount, see projects.py's
    get_estimate_items_for_invoice) would silently go negative, and its
    invoiced_pct could read over 100%, with nothing anywhere stopping it or
    explaining why. Mirrors invoices.py's own _validate_invoice_amounts,
    just checked from the estimate side of the same relationship."""
    invoiced_rows = await db_get("invoice_line_items", f"?source_line_item_id=eq.{item_id}&select=amount")
    already_invoiced = sum(r.get("amount") or 0 for r in invoiced_rows)
    if already_invoiced and round(new_owner_price, 2) < round(already_invoiced, 2):
        raise HTTPException(
            status_code=400,
            detail=(
                f"Can't lower this line item's price below ${already_invoiced:,.2f} -- "
                f"that much of it has already been invoiced."
            ),
        )


async def _recalc_estimate_totals(estimate_id: str) -> None:
    items = await db_get("estimate_line_items", f"?estimate_id=eq.{estimate_id}")
    pm_fee_total = sum(i.get("owner_price") or 0 for i in items if i.get("bucket") == "pm_fee")
    construction_total = sum(i.get("owner_price") or 0 for i in items if i.get("bucket") == "construction")
    allowance_total = sum(i.get("owner_price") or 0 for i in items if i.get("bucket") == "allowance")
    await db_patch(
        "estimates",
        estimate_id,
        {
            "pm_fee_total": round(pm_fee_total, 2),
            "construction_total_owner_price": round(construction_total, 2),
            "allowance_total": round(allowance_total, 2),
            "grand_total_owner_price": round(pm_fee_total + construction_total + allowance_total, 2),
        },
    )


@router.get("", response_model=list[EstimateOut])
async def list_estimates(project_id: Optional[str] = None, _: CurrentUser = Depends(get_current_user)):
    # Embeds the project's status (not just its name) so the Estimates page
    # can group estimates by where their project actually is in the
    # pipeline (active / pre construction / closed / etc.) instead of only
    # showing the estimate's own draft/sent/approved status.
    query = "?order=created_at.desc&select=*,projects(name,status)"
    if project_id:
        query += f"&project_id=eq.{project_id}"
    return await db_get("estimates", query)


@router.get("/line-items/search", response_model=list[LineItemReference])
async def search_line_items(
    cost_code_id: Optional[str] = None,
    q: Optional[str] = None,
    exclude_estimate_id: Optional[str] = None,
    _: CurrentUser = Depends(get_current_user),
):
    if not cost_code_id and not q:
        return []
    query = "?order=created_at.desc&limit=25&select=*,estimates(project_id,projects(name))"
    if cost_code_id:
        query += f"&cost_code_id=eq.{cost_code_id}"
    if q:
        escaped = q.replace(",", "").replace("(", "").replace(")", "")
        query += f"&or=(title.ilike.*{escaped}*,description.ilike.*{escaped}*)"
    if exclude_estimate_id:
        query += f"&estimate_id=neq.{exclude_estimate_id}"
    rows = await db_get("estimate_line_items", query)
    results = []
    for r in rows:
        est = r.get("estimates") or {}
        proj = est.get("projects") or {}
        results.append(
            LineItemReference(
                id=r["id"],
                estimate_id=r["estimate_id"],
                project_name=(proj.get("name") or "").split("|")[0].strip() or None,
                title=r["title"],
                description=r.get("description"),
                quantity=r["quantity"],
                unit=r.get("unit"),
                unit_cost=r["unit_cost"],
                cost_type=r["cost_type"],
                builder_cost=r["builder_cost"],
                markup_type=r["markup_type"],
                markup_value=r["markup_value"],
                owner_price=r["owner_price"],
                estimated_days=r.get("estimated_days"),
                notes_internal=r.get("notes_internal"),
                notes_external=r.get("notes_external"),
                created_at=r["created_at"],
            )
        )
    return results


@router.get("/{estimate_id}", response_model=EstimateOut)
async def get_estimate(estimate_id: str, _: CurrentUser = Depends(get_current_user)):
    rows = await db_get("estimates", f"?id=eq.{estimate_id}&select=*,projects(name)")
    if not rows:
        raise HTTPException(status_code=404, detail="Estimate not found")
    return rows[0]


@router.post("", response_model=EstimateOut)
async def create_estimate(body: EstimateCreate, _: CurrentUser = Depends(get_current_user)):
    data = body.model_dump(exclude_none=True)
    if "closing_text" not in data or "introductory_text" not in data:
        defaults = await get_or_create_estimate_text_defaults()
        data.setdefault("closing_text", defaults.get("closing_text") or DEFAULT_CLOSING_TEXT)
        data.setdefault("introductory_text", defaults.get("introductory_text"))
    rows = await db_post("estimates", data)
    estimate = rows[0]
    if body.pm_fee_total > 0:
        await db_post(
            "estimate_line_items",
            {
                "estimate_id": estimate["id"],
                "bucket": "pm_fee",
                "group_name": "PM Fee",
                "title": "Project management fee",
                "quantity": 1,
                "unit_cost": body.pm_fee_total,
                "cost_type": "none",
                "markup_type": "flat",
                "markup_value": 0,
                "builder_cost": body.pm_fee_total,
                "owner_price": body.pm_fee_total,
                "sort_order": 1,
            },
        )
        await _recalc_estimate_totals(estimate["id"])
        full = await db_get("estimates", f"?id=eq.{estimate['id']}&select=*,projects(name)")
        return full[0]
    return estimate


@router.patch("/{estimate_id}", response_model=EstimateOut)
async def update_estimate(estimate_id: str, body: EstimateUpdate, _: CurrentUser = Depends(get_current_user)):
    current = await db_get("estimates", f"?id=eq.{estimate_id}&select=status")
    if not current:
        raise HTTPException(status_code=404, detail="Estimate not found")
    old_status = current[0]["status"]

    # exclude_unset (not exclude_none) -- a caller may need to explicitly
    # clear a field (e.g. removing an approval_deadline or closing_text),
    # and that null has to reach the database instead of being silently
    # dropped. Same fix already made for clients/projects/invoices/etc.
    await db_patch("estimates", estimate_id, body.model_dump(exclude_unset=True))
    full = await db_get("estimates", f"?id=eq.{estimate_id}&select=*,projects(name)")
    if not full:
        raise HTTPException(status_code=404, detail="Estimate not found")
    estimate = full[0]

    # Approving an estimate is the moment the client's actual signed total
    # becomes known -- sync projects.contract_value to it (plus whatever
    # change orders are already approved, the same additive formula
    # financial-summary itself uses) so the two "what's this contract
    # worth" numbers in this app can't silently drift apart. Mirrors the
    # exact reasoning change_orders.py already applies on CO approval.
    # One-directional on purpose: un-approving an estimate doesn't try to
    # guess what to revert contract_value to.
    new_status = body.status if body.status is not None else old_status
    if new_status == "approved" and old_status != "approved":
        approved_cos = await db_get(
            "change_orders", f"?project_id=eq.{estimate['project_id']}&status=eq.approved&select=owner_price"
        )
        co_total = sum(c.get("owner_price") or 0 for c in approved_cos)
        new_contract_value = round((estimate.get("grand_total_owner_price") or 0) + co_total, 2)
        await db_patch("projects", estimate["project_id"], {"contract_value": new_contract_value})

    return estimate


@router.post("/{estimate_id}/duplicate", response_model=EstimateOut)
async def duplicate_estimate(estimate_id: str, _: CurrentUser = Depends(get_current_user)):
    originals = await db_get("estimates", f"?id=eq.{estimate_id}")
    if not originals:
        raise HTTPException(status_code=404, detail="Estimate not found")
    original = originals[0]
    siblings = await db_get(
        "estimates", f"?project_id=eq.{original['project_id']}&select=version&order=version.desc&limit=1"
    )
    next_version = (siblings[0]["version"] + 1) if siblings else 1

    new_estimate = (
        await db_post(
            "estimates",
            {
                "project_id": original["project_id"],
                "version": next_version,
                "status": "draft",
                "title": original.get("title"),
                "notes_internal": original.get("notes_internal"),
                "approval_deadline": original.get("approval_deadline"),
                "introductory_text": original.get("introductory_text"),
                "closing_text": original.get("closing_text") or DEFAULT_CLOSING_TEXT,
            },
        )
    )[0]

    items = await db_get("estimate_line_items", f"?estimate_id=eq.{estimate_id}&order=sort_order.asc")
    if items:
        await db_post_many(
            "estimate_line_items",
            [
                {
                    "estimate_id": new_estimate["id"],
                    "cost_code_id": item.get("cost_code_id"),
                    "group_name": item.get("group_name"),
                    "bucket": item.get("bucket"),
                    "title": item.get("title"),
                    "description": item.get("description"),
                    "quantity": item.get("quantity"),
                    "unit": item.get("unit"),
                    "unit_cost": item.get("unit_cost"),
                    "cost_type": item.get("cost_type"),
                    "builder_cost": item.get("builder_cost"),
                    "markup_type": item.get("markup_type"),
                    "markup_value": item.get("markup_value"),
                    "owner_price": item.get("owner_price"),
                    "notes_internal": item.get("notes_internal"),
                    "notes_external": item.get("notes_external"),
                    "sort_order": item.get("sort_order"),
                }
                for item in items
            ],
        )
    await _recalc_estimate_totals(new_estimate["id"])
    full = await db_get("estimates", f"?id=eq.{new_estimate['id']}&select=*,projects(name)")
    return full[0]


@router.get("/{estimate_id}/items", response_model=list[LineItemOut])
async def list_line_items(estimate_id: str, _: CurrentUser = Depends(get_current_user)):
    return await db_get(
        "estimate_line_items", f"?estimate_id=eq.{estimate_id}&order=sort_order.asc&select=*,cost_codes(code,name)"
    )


@router.post("/{estimate_id}/items", response_model=LineItemOut)
async def create_line_item(estimate_id: str, body: LineItemCreate, _: CurrentUser = Depends(get_current_user)):
    builder_cost, owner_price = _compute_costs(body.quantity, body.unit_cost, body.markup_type, body.markup_value)
    data = {
        **body.model_dump(exclude_none=True),
        "estimate_id": estimate_id,
        "builder_cost": builder_cost,
        "owner_price": owner_price,
    }
    rows = await db_post("estimate_line_items", data)
    await _recalc_estimate_totals(estimate_id)
    full = await db_get("estimate_line_items", f"?id=eq.{rows[0]['id']}&select=*,cost_codes(code,name)")
    return full[0]


@router.patch("/{estimate_id}/items/{item_id}", response_model=LineItemOut)
async def update_line_item(
    estimate_id: str, item_id: str, body: LineItemUpdate, _: CurrentUser = Depends(get_current_user)
):
    existing_rows = await db_get("estimate_line_items", f"?id=eq.{item_id}")
    if not existing_rows:
        raise HTTPException(status_code=404, detail="Line item not found")
    existing = existing_rows[0]
    # exclude_unset (not exclude_none) -- a caller may need to explicitly
    # clear a field (e.g. removing a cost_code_id or notes_external), and
    # that null has to reach the database instead of being silently dropped.
    updates = body.model_dump(exclude_unset=True)
    merged = {**existing, **updates}
    builder_cost, owner_price = _compute_costs(
        merged.get("quantity") or 0, merged.get("unit_cost") or 0, merged.get("markup_type") or "percent", merged.get("markup_value") or 0
    )
    await _check_not_below_invoiced(item_id, owner_price)
    updates["builder_cost"] = builder_cost
    updates["owner_price"] = owner_price
    await db_patch("estimate_line_items", item_id, updates)
    await _recalc_estimate_totals(estimate_id)
    full = await db_get("estimate_line_items", f"?id=eq.{item_id}&select=*,cost_codes(code,name)")
    return full[0]


@router.delete("/{estimate_id}/items/{item_id}")
async def delete_line_item(estimate_id: str, item_id: str, _: CurrentUser = Depends(get_current_user)):
    await db_delete("estimate_line_items", item_id)
    await _recalc_estimate_totals(estimate_id)
    return {"ok": True}


GROUP_LABEL_FALLBACK = "Ungrouped"


async def _gather_export_data(estimate_id: str):
    estimates = await db_get("estimates", f"?id=eq.{estimate_id}&select=*,projects(name,address,clients(first_name,last_name))")
    if not estimates:
        raise HTTPException(status_code=404, detail="Estimate not found")
    estimate = estimates[0]
    items = await db_get(
        "estimate_line_items", f"?estimate_id=eq.{estimate_id}&order=sort_order.asc&select=*,cost_codes(code,name)"
    )
    groups: dict[str, list[dict]] = {}
    for item in items:
        key = item.get("group_name") or GROUP_LABEL_FALLBACK
        groups.setdefault(key, []).append(item)
    return estimate, groups


class _NumberedCanvas(Canvas):
    """Standard reportlab two-pass trick for "Page N of M" -- the total page
    count isn't known until the whole document has been laid out, so each
    page's canvas state is buffered via showPage() and only actually drawn
    (with the number stamped on) once save() knows the final count."""

    def __init__(self, *args, **kwargs):
        Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_page_number(num_pages)
            Canvas.showPage(self)
        Canvas.save(self)

    def _draw_page_number(self, page_count):
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.grey)
        self.drawRightString(letter[0] - 0.5 * inch, 0.3 * inch, f"Page {self._pageNumber} of {page_count}")


@router.get("/{estimate_id}/export/pdf")
async def export_estimate_pdf(estimate_id: str, _: CurrentUser = Depends(get_current_user)):
    estimate, groups = await _gather_export_data(estimate_id)
    project = estimate.get("projects") or {}
    client = (project.get("clients") or {}) if project else {}
    client_name = f"{client.get('first_name') or ''} {client.get('last_name') or ''}".strip()
    project_name = (project.get("name") or "").split("|")[0].strip()

    who = client_name or project_name
    breadcrumb = who + (f" | {project_name}" if project_name and project_name != who else "")
    # Every free-text value below (client/project name, estimate title, group
    # names, item titles/descriptions/units) is user-entered and now flows
    # through reportlab's Paragraph(), which parses a small XML-like markup
    # subset -- an unescaped '<' followed by a letter with no matching '>'
    # (e.g. someone typing "a<b" or "<8ft" as a quantity note) throws a
    # paraparser syntax error and 500s the whole export. Escaping & / < / >
    # up front, before any Paragraph() call, closes that off entirely.
    breadcrumb = _xml_escape(breadcrumb)

    buf = io.BytesIO()
    # Slightly tighter side margins than reportlab's 1in default (0.6in,
    # matching the density of the BuilderTrend reference proposal) --
    # PAGE_WIDTH is the one number every table's colWidths below is sized
    # against, so a table summing to less than this leaves a lopsided gap on
    # the right, and one summing to more silently overflows the margin. Both
    # happened before this pass (the two-column breadcrumb/print-date row
    # summed to 7.2in against a 6.5in page -- overflowing the right margin --
    # while the line-item table summed to only 6.4in, leaving a stray gap).
    side_margin = 0.6 * inch
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch, leftMargin=side_margin, rightMargin=side_margin)
    PAGE_WIDTH = letter[0] - 2 * side_margin
    styles = getSampleStyleSheet()
    wordmark_h1 = ParagraphStyle("wordmark_h1", parent=styles["Heading1"], fontSize=14, alignment=1, spaceBefore=4, spaceAfter=1)
    company_line = ParagraphStyle("company_line", parent=styles["Normal"], fontSize=8, alignment=1, textColor=colors.grey, spaceAfter=10)
    group_header = ParagraphStyle("group_header", parent=styles["Normal"], fontSize=10, fontName="Helvetica-Bold", textColor=colors.white)
    group_subtotal = ParagraphStyle("group_subtotal", parent=group_header, alignment=2)
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=8.5, leading=12)
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8.2, leading=11)
    cell_right = ParagraphStyle("cell_right", parent=cell, alignment=2)
    th = ParagraphStyle("th", parent=cell, fontName="Helvetica-Bold", textColor=branding.BRAND_BROWN)
    th_right = ParagraphStyle("th_right", parent=th, alignment=2)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, textColor=colors.grey)
    small_right = ParagraphStyle("small_right", parent=small, alignment=2)
    title_style = ParagraphStyle("title", parent=styles["Normal"], fontSize=13, spaceBefore=6, spaceAfter=2, fontName="Helvetica-Bold")

    # Centered logo + wordmark + company contact line -- matches the
    # BuilderTrend reference proposal's header layout.
    elements = [
        Image(branding.LOGO_PATH, width=0.55 * inch, height=0.55 * inch, hAlign="CENTER"),
        Paragraph("Mud &amp; Marble", wordmark_h1),
        Paragraph(branding.COMPANY_ADDRESS_LINE, company_line),
    ]

    # Left: who this proposal is for. Right: print date. Same row, small text --
    # matches the reference's breadcrumb + print-date line above the title.
    print_date = datetime.now()
    header_row = Table(
        [[
            Paragraph(breadcrumb, small),
            Paragraph(f"Print Date: {print_date.month}-{print_date.day}-{print_date.year}", small_right),
        ]],
        colWidths=[PAGE_WIDTH * 0.6, PAGE_WIDTH * 0.4],
    )
    header_row.setStyle(
        TableStyle(
            [
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    elements.append(header_row)
    title_text = _xml_escape(estimate.get("title")) if estimate.get("title") else f"Proposal for {breadcrumb}"
    elements.append(Paragraph(title_text, title_style))
    elements.append(HRFlowable(width="100%", thickness=0.75, color=colors.lightgrey, spaceAfter=8))

    intro = estimate.get("introductory_text")
    if intro:
        elements.append(Paragraph(rich_text_to_pdf_markup(intro), body))
        elements.append(Spacer(1, 8))

    # Column widths sum to exactly PAGE_WIDTH -- Item/Description get the
    # bulk of the space (that's the text that actually wraps), the three
    # numeric columns are just wide enough for "$12,345.67" without wrapping.
    item_col = PAGE_WIDTH * 0.20
    desc_col = PAGE_WIDTH * 0.38
    qty_col = PAGE_WIDTH * 0.14
    unit_price_col = PAGE_WIDTH * 0.14
    price_col = PAGE_WIDTH - item_col - desc_col - qty_col - unit_price_col
    col_widths = [item_col, desc_col, qty_col, unit_price_col, price_col]

    for group_name, items in groups.items():
        group_subtotal_value = sum(item.get("owner_price") or 0 for item in items)
        # A shaded, full-width band per group (title left, subtotal right) --
        # gives each section clear visual separation and a running subtotal,
        # instead of a plain bold heading floating with no boundary, and
        # reuses the same brand-brown/cream pairing as the item table's own
        # header row so the two read as one cohesive block, not two
        # differently-styled pieces stacked on top of each other.
        group_band = Table(
            [[Paragraph(_xml_escape(group_name.upper()), group_header), Paragraph(f"${group_subtotal_value:,.2f}", group_subtotal)]],
            colWidths=[PAGE_WIDTH * 0.7, PAGE_WIDTH * 0.3],
        )
        group_band.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), branding.BRAND_BROWN),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (0, 0), 6),
                    ("RIGHTPADDING", (-1, 0), (-1, 0), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elements.append(group_band)

        table_data = [
            [
                Paragraph("Item", th),
                Paragraph("Description", th),
                Paragraph("Qty/Unit", th_right),
                Paragraph("Unit Price", th_right),
                Paragraph("Price", th_right),
            ]
        ]
        for item in items:
            # Cost codes are internal categorization, not something a client
            # needs to see on their proposal -- title only.
            item_label = _xml_escape(item.get("title") or "")
            qty_unit = _xml_escape(f"{item.get('quantity') or 0:g}" + (f" {item['unit']}" if item.get("unit") else ""))
            # Client-facing figures only -- unit_cost/builder_cost are internal
            # margin data and must never appear on anything a client sees.
            # "Unit Price" here is a derived per-unit client price
            # (owner_price / quantity), the same thing BuilderTrend's export
            # shows, not the builder's cost.
            qty = item.get("quantity") or 0
            owner_price = item.get("owner_price") or 0
            client_unit_price = (owner_price / qty) if qty else owner_price
            table_data.append(
                [
                    Paragraph(item_label, cell),
                    Paragraph(_xml_escape(item.get("description") or ""), cell),
                    Paragraph(qty_unit, cell_right),
                    Paragraph(f"${client_unit_price:,.2f}", cell_right),
                    Paragraph(f"${owner_price:,.2f}", cell_right),
                ]
            )
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), branding.BRAND_CREAM),
                    ("LINEBELOW", (0, 0), (-1, 0), 0.75, branding.BRAND_BROWN),
                    ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.lightgrey),
                    # Subtle zebra striping on data rows only (row 0 is the
                    # header, already shaded cream above) -- easier to track
                    # a row across five columns than a flat white table.
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAF8F3")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (0, -1), 6),
                    ("RIGHTPADDING", (-1, 0), (-1, -1), 6),
                ]
            )
        )
        elements.append(t)
        elements.append(Spacer(1, 12))

    total = estimate.get("grand_total_owner_price") or 0
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"<b>Total Price: ${total:,.2f}</b>", ParagraphStyle("total", parent=styles["Normal"], fontSize=12, alignment=2)))
    elements.append(Spacer(1, 16))

    closing = estimate.get("closing_text") or DEFAULT_CLOSING_TEXT
    elements.append(Paragraph(rich_text_to_pdf_markup(closing), body))
    elements.append(Spacer(1, 6))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Signature: _______________________________________", body))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Date: _______________________________________", body))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Print Name: _______________________________________", body))

    doc.build(elements, canvasmaker=_NumberedCanvas)
    pdf_bytes = buf.getvalue()
    buf.close()

    filename = f"proposal-{project_name or 'estimate'}-v{estimate['version']}.pdf".replace(" ", "-")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{estimate_id}/export/excel")
async def export_estimate_excel(estimate_id: str, _: CurrentUser = Depends(get_current_user)):
    estimate, groups = await _gather_export_data(estimate_id)
    project = estimate.get("projects") or {}
    project_name = (project.get("name") or "").split("|")[0].strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Proposal"
    header_font = Font(bold=True)

    ws.append([estimate.get("title") or f"Proposal for {project_name}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    for group_name, items in groups.items():
        ws.append([group_name])
        ws.cell(row=ws.max_row, column=1).font = header_font
        ws.append(["Item", "Description", "Qty", "Unit", "Unit Price", "Price"])
        for cell in ws[ws.max_row]:
            cell.font = header_font
        for item in items:
            # Client-facing figures only -- unit_cost/builder_cost are internal
            # margin data and must never appear in anything exported for a
            # client. "Unit Price" is a derived client per-unit price
            # (owner_price / quantity), not the builder's cost.
            qty = item.get("quantity") or 0
            owner_price = item.get("owner_price") or 0
            client_unit_price = (owner_price / qty) if qty else owner_price
            ws.append(
                [
                    item.get("title"),
                    item.get("description"),
                    item.get("quantity"),
                    item.get("unit"),
                    client_unit_price,
                    owner_price,
                ]
            )
        ws.append([])

    ws.append(["", "", "", "", "Total", estimate.get("grand_total_owner_price") or 0])
    ws.cell(row=ws.max_row, column=5).font = header_font
    ws.cell(row=ws.max_row, column=6).font = header_font

    for col, width in zip("ABCDEF", [28, 40, 8, 8, 12, 12]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()
    buf.close()

    filename = f"proposal-{project_name or 'estimate'}-v{estimate['version']}.xlsx".replace(" ", "-")
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
