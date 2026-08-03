import asyncio
import io
import uuid

import openpyxl
import xlrd
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile

from ..ai_provider import (
    ProviderError,
    build_scan_content_block,
    extract_estimate_from_scan,
    extract_invoice_from_document,
    extract_transactions_from_scan,
)
from ..deps import CurrentUser, get_current_user
from ..inhouse_import import (
    diff_estimate_scan_items,
    diff_fields_helper,
    diff_transaction_scan_items,
    parse_contractors_sheet,
    parse_estimate_sheet,
    parse_quickbooks_sheet,
)
from ..schemas.job_import import (
    ContractorBlock,
    EstimateSheetPreview,
    EstimateSheetRow,
    InHouseSheetPreview,
    InvoiceScanPreview,
    InvoiceScanRow,
    JobImportStatus,
    TransactionSheetRow,
)
from ..storage_client import upload_object
from ..supabase_client import db_get, db_post

router = APIRouter(prefix="/job-import", tags=["job-import"])

FILES_BUCKET = "project-files"

EXCEL_EXTS = (".xlsx", ".xlsm", ".xls")
SCAN_EXTS = (".pdf", ".jpg", ".jpeg", ".png")


def _ext(filename: str) -> str:
    return ("." + filename.rsplit(".", 1)[-1].lower()) if filename and "." in filename else ""


def _load_workbook(content: bytes, filename: str):
    """openpyxl only understands the zip-based .xlsx/.xlsm container -- a legacy
    .xls file is a different, non-zip binary format and openpyxl raises "File is
    not a zip file" if handed one directly. For .xls, read it with xlrd and
    reconstruct an equivalent in-memory openpyxl Workbook so every downstream
    parser (which expects openpyxl's Worksheet/cell API) works unchanged."""
    if _ext(filename) == ".xls":
        return _xls_to_openpyxl(content)
    return openpyxl.load_workbook(io.BytesIO(content), data_only=True)


def _xls_to_openpyxl(content: bytes):
    src = xlrd.open_workbook(file_contents=content)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in src.sheet_names():
        sheet = src.sheet_by_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name[:31])
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(value, src.datemode)
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    value = None
                ws.cell(row=r + 1, column=c + 1, value=value)
    return wb


async def _notify_conflicts(user_id: str, project_id: str, project_name: str, count: int, section: str) -> None:
    """Surfaces a review prompt in the recipient's notification bell whenever a
    re-imported row doesn't match what's already on record -- Shannon (or
    whoever's running the import) gets a durable pointer back to the wizard
    even if she doesn't resolve it in the moment."""
    if count <= 0:
        return
    await db_post(
        "notifications",
        {
            "user_id": user_id,
            "type": "job_import_conflict",
            "source_type": "job_import",
            "source_id": project_id,
            "project_id": project_id,
            "message": (
                f"{project_name}: {count} {section} row(s) from this import don't match what's already "
                f"on file. Which values are correct?"
            ),
        },
    )


def _find_sheet(wb, name: str):
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == name.lower():
            return wb[sheet_name]
    return None


@router.get("/status", response_model=list[JobImportStatus])
async def get_status(_: CurrentUser = Depends(get_current_user)):
    (
        projects,
        estimates,
        invoices,
        change_orders,
        transactions,
        sub_items,
    ) = await asyncio.gather(
        db_get("projects", "?is_archived=eq.false&select=id,name"),
        db_get("estimates", "?select=project_id,grand_total_owner_price"),
        db_get("invoices", "?select=project_id"),
        db_get("change_orders", "?select=project_id"),
        db_get("transactions", "?select=project_id"),
        db_get("project_subcontractor_items", "?select=project_id"),
    )

    has_estimate = {e["project_id"] for e in estimates if (e.get("grand_total_owner_price") or 0) > 0}
    has_financials = {i["project_id"] for i in invoices} | {c["project_id"] for c in change_orders}
    has_inhouse = {t["project_id"] for t in transactions} | {s["project_id"] for s in sub_items}

    return [
        JobImportStatus(
            project_id=p["id"],
            project_name=p["name"],
            has_estimate=p["id"] in has_estimate,
            has_financials=p["id"] in has_financials,
            has_inhouse=p["id"] in has_inhouse,
        )
        for p in projects
    ]


@router.post("/{project_id}/estimate-sheet/preview", response_model=EstimateSheetPreview)
async def preview_estimate_sheet(
    project_id: str, file: UploadFile = File(...), current_user: CurrentUser = Depends(get_current_user)
):
    content = await file.read()
    ext = _ext(file.filename or "")
    if ext not in EXCEL_EXTS and ext not in SCAN_EXTS:
        raise HTTPException(status_code=400, detail="Unsupported file type -- use an Excel workbook, JPEG, PNG, or PDF")

    project_rows, existing_estimates, cost_codes = await asyncio.gather(
        db_get("projects", f"?id=eq.{project_id}&select=name"),
        db_get("estimates", f"?project_id=eq.{project_id}&order=version.desc&limit=1"),
        db_get("cost_codes", "?is_active=eq.true&select=id,code"),
    )
    if not project_rows:
        raise HTTPException(status_code=404, detail="Project not found")
    project_name = project_rows[0]["name"].split("|")[0].strip()
    codes_by_code = {c["code"].strip().lower(): c["id"] for c in cost_codes if c.get("code")}

    existing_estimate_id = None
    existing_by_key: dict = {}
    if existing_estimates:
        existing_estimate_id = existing_estimates[0]["id"]
        existing_items = await db_get(
            "estimate_line_items",
            f"?estimate_id=eq.{existing_estimate_id}&select=id,title,unit_cost,quantity,markup_type,markup_value,description,notes_internal,bucket",
        )
        existing_by_key = {(i["title"], i["unit_cost"]): i for i in existing_items}

    dropped_count = 0
    if ext in EXCEL_EXTS:
        try:
            wb = _load_workbook(content, file.filename or "")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read that file: {e}") from e
        ws = _find_sheet(wb, "Estimate")
        if ws is None:
            raise HTTPException(status_code=400, detail="No sheet named 'Estimate' found in that file")
        parsed = parse_estimate_sheet(ws, existing_by_key)
    else:
        try:
            block = build_scan_content_block(content, file.filename or "", file.content_type)
            items, dropped_count = await extract_estimate_from_scan(block)
        except ProviderError as e:
            raise HTTPException(status_code=502, detail=str(e)) from e
        parsed = diff_estimate_scan_items(items, existing_by_key)

    rows = []
    conflict_count = 0
    for row in parsed:
        code = row.get("cost_code")
        if row.get("conflict"):
            conflict_count += 1
        rows.append(
            EstimateSheetRow(
                **row,
                matched_cost_code_id=codes_by_code.get(code.lower()) if code else None,
            )
        )
    await _notify_conflicts(current_user.id, project_id, project_name, conflict_count, "estimate line item")
    return EstimateSheetPreview(rows=rows, existing_estimate_id=existing_estimate_id, dropped_count=dropped_count)


@router.post("/{project_id}/inhouse-sheet/preview", response_model=InHouseSheetPreview)
async def preview_inhouse_sheet(
    project_id: str, file: UploadFile = File(...), current_user: CurrentUser = Depends(get_current_user)
):
    content = await file.read()
    ext = _ext(file.filename or "")
    if ext not in EXCEL_EXTS and ext not in SCAN_EXTS:
        raise HTTPException(status_code=400, detail="Unsupported file type -- use an Excel workbook, JPEG, PNG, or PDF")

    project_rows, existing_tx, cost_codes, subcontractors, existing_sub_items = await asyncio.gather(
        db_get("projects", f"?id=eq.{project_id}&select=name"),
        db_get(
            "transactions",
            f"?project_id=eq.{project_id}&select=id,transaction_date,amount,description,vendor,transaction_type,payment_source",
        ),
        db_get("cost_codes", "?is_active=eq.true&select=id,code"),
        db_get("subcontractors", "?select=id,company_name"),
        db_get("project_subcontractor_items", f"?project_id=eq.{project_id}&select=id,subcontractor_id,description,amount"),
    )
    if not project_rows:
        raise HTTPException(status_code=404, detail="Project not found")
    project_name = project_rows[0]["name"].split("|")[0].strip()
    codes_by_code = {c["code"].strip().lower(): c["id"] for c in cost_codes if c.get("code")}
    subs_by_name = {s["company_name"].strip().lower(): s["id"] for s in subcontractors if s.get("company_name")}
    existing_tx_by_key = {
        (t["transaction_date"][:10], round(t["amount"], 2), (t.get("description") or "")[:60]): t for t in existing_tx
    }
    existing_tx_keys = set(existing_tx_by_key.keys())

    existing_items_by_sub: dict = {}
    for item in existing_sub_items:
        existing_items_by_sub.setdefault(item["subcontractor_id"], {})[item["description"]] = item

    transactions: list[TransactionSheetRow] = []
    contractors: list[ContractorBlock] = []
    conflict_count = 0
    dropped_count = 0

    if ext in EXCEL_EXTS:
        try:
            wb = _load_workbook(content, file.filename or "")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read that file: {e}") from e
        qb_ws = _find_sheet(wb, "Quickbooks")
        contractors_ws = _find_sheet(wb, "Contractors")
        if qb_ws is None and contractors_ws is None:
            raise HTTPException(
                status_code=400, detail="No sheet named 'Quickbooks' or 'Contractors' found in that file"
            )
        if qb_ws is not None:
            parsed_tx = parse_quickbooks_sheet(qb_ws, existing_tx_by_key)
            for row in parsed_tx:
                code = row.get("cost_code")
                if row.get("conflict"):
                    conflict_count += 1
                transactions.append(
                    TransactionSheetRow(**row, matched_cost_code_id=codes_by_code.get(code.lower()) if code else None)
                )
        if contractors_ws is not None:
            parsed_blocks = parse_contractors_sheet(contractors_ws, subs_by_name, existing_items_by_sub, existing_tx_keys)
            for block in parsed_blocks:
                contractors.append(ContractorBlock(**block))
                conflict_count += sum(1 for item in block["contract_items"] if item.get("conflict"))
    else:
        # A photographed/scanned in-house sheet realistically only captures a
        # bank statement or receipt page -- transactions only. The
        # "Contractors" sheet (subcontractor blocks with contract items +
        # payments) stays Excel-only; `contractors` is intentionally left empty.
        try:
            block = build_scan_content_block(content, file.filename or "", file.content_type)
            items, dropped_count = await extract_transactions_from_scan(block)
        except ProviderError as e:
            raise HTTPException(status_code=502, detail=str(e)) from e
        parsed_tx = diff_transaction_scan_items(items, existing_tx_by_key)
        for row in parsed_tx:
            code = row.get("cost_code")
            if row.get("conflict"):
                conflict_count += 1
            transactions.append(
                TransactionSheetRow(**row, matched_cost_code_id=codes_by_code.get(code.lower()) if code else None)
            )

    await _notify_conflicts(current_user.id, project_id, project_name, conflict_count, "in-house sheet")
    return InHouseSheetPreview(transactions=transactions, contractors=contractors, dropped_count=dropped_count)


def _flatten_workbook_to_text(content: bytes, filename: str = "") -> str:
    """No fixed invoice template exists (unlike Estimate/In-House, which mirror
    Brent's one real template), so an Excel invoice doesn't get a deterministic
    parser -- every cell gets dumped into a plain-text grid and handed to the
    same AI extraction function used for images. Claude reads messy tabular
    text fine."""
    wb = _load_workbook(content, filename)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows():
            values = [str(c.value) for c in row if c.value not in (None, "")]
            if values:
                lines.append(" | ".join(values))
    return "\n".join(lines)


async def _archive_scan(project_id: str, user_id: str, filename: str, content: bytes, content_type: str, label: str) -> None:
    """Best-effort, non-blocking: saves the original scan into the project's
    Files tab for later reference. A failed archive copy must never block the
    actual import -- this runs after a successful preview and swallows its
    own errors."""
    try:
        storage_path = f"{project_id}/{uuid.uuid4()}_{filename or 'scan'}"
        await upload_object(FILES_BUCKET, storage_path, content, content_type or "application/octet-stream")
        await db_post(
            "project_files",
            {
                "project_id": project_id,
                "uploaded_by": user_id,
                "file_name": f"{label} - {filename or 'scan'}",
                "file_type": "other",
                "mime_type": content_type,
                "size_bytes": len(content),
                "storage_path": storage_path,
            },
        )
    except Exception:
        pass


_INVOICE_DIFF_FIELDS = [
    ("Type", "invoice_type", "invoice_type"),
    ("Amount due", "amount_due", "amount_due"),
    ("Due date", "due_date", "due_date"),
    ("Notes", "notes_external", "notes_external"),
]


@router.post("/{project_id}/invoice-scan/preview", response_model=InvoiceScanPreview)
async def preview_invoice_scan(
    project_id: str, file: UploadFile = File(...), current_user: CurrentUser = Depends(get_current_user)
):
    content = await file.read()
    filename = file.filename or ""
    ext = _ext(filename)
    if ext not in EXCEL_EXTS and ext not in SCAN_EXTS:
        raise HTTPException(status_code=400, detail="Unsupported file type -- use an Excel workbook, JPEG, PNG, or PDF")

    project_rows, existing_invoices = await asyncio.gather(
        db_get("projects", f"?id=eq.{project_id}&select=name"),
        db_get("invoices", f"?project_id=eq.{project_id}&select=id,invoice_number,invoice_type,amount_due,due_date,notes_external"),
    )
    if not project_rows:
        raise HTTPException(status_code=404, detail="Project not found")
    project_name = project_rows[0]["name"].split("|")[0].strip()
    existing_by_number = {i["invoice_number"]: i for i in existing_invoices if i.get("invoice_number")}

    try:
        if ext in EXCEL_EXTS:
            content_block = {"type": "text", "text": _flatten_workbook_to_text(content, filename)}
        else:
            content_block = build_scan_content_block(content, filename, file.content_type)
        extraction = await extract_invoice_from_document(content_block)
    except ProviderError as e:
        raise HTTPException(status_code=502, detail=str(e)) from e
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read that file: {e}") from e

    # No reliable fallback key exists when a scan doesn't show an invoice
    # number -- rather than risk a silent wrong match on amount/date alone,
    # an invoice with no number is always treated as new and left to the
    # reviewer to notice if it's actually a duplicate.
    existing = existing_by_number.get(extraction.invoice_number) if extraction.invoice_number else None
    incoming = {
        "invoice_type": extraction.invoice_type,
        "amount_due": extraction.amount_due,
        "due_date": extraction.due_date,
        "notes_external": extraction.notes_external,
    }
    diff = diff_fields_helper(existing, incoming, _INVOICE_DIFF_FIELDS) if existing else []

    row = InvoiceScanRow(
        invoice_number=extraction.invoice_number,
        **incoming,
        confidence=extraction.confidence,
        uncertain_fields=extraction.uncertain_fields,
        already_present=existing is not None,
        existing_id=existing["id"] if existing else None,
        conflict=bool(diff),
        diff=diff,
    )
    await _notify_conflicts(current_user.id, project_id, project_name, 1 if diff else 0, "invoice")
    await _archive_scan(project_id, current_user.id, filename, content, file.content_type or "", "Invoice scan")
    return InvoiceScanPreview(row=row)
